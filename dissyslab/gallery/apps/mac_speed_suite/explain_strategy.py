#!/usr/bin/env python3
"""Show a strategy's working, bar by bar, in Excel.

Why this exists
---------------
`report.html` answers "how did it do". This answers a different
question: **is this the strategy I meant?** A tester who thinks about
markets, not about Python, cannot check that from a signal column. They
need the intermediate quantities -- the channel bounds, the two moving
averages -- and the rule that turned them into a position.

Each derived quantity appears twice: the value the Python produced, and
the same quantity as a live Excel formula over the price cells. The
formula is there to be *read*, and edited. `=MAX(D7:D9)` in row 10 says
"the three rows above this one, not this one" without anyone explaining
the boundary convention, and if that is not your rule you can change it
and watch the signal column move.

The two columns do not verify each other. Both express one author's
understanding of the rule, so a misreading appears in both. What they
catch is an inconsistency between the two expressions -- and what the
formula gives you is a specification you can read.

Usage
-----
    python3 explain_strategy.py --strategy donchian --variant 20
    python3 explain_strategy.py --strategy mac --variant med --ticker NVDA
    python3 explain_strategy.py --strategy donchian --strategy mac --rows 25
    python3 explain_strategy.py --strategy donchian --bars 300:340
    python3 explain_strategy.py --strategy turtle --variant s1
    python3 explain_strategy.py --strategy rs --variant fast --peers AMD,NVDA,TSLA

With no --bars, a window is chosen that contains a signal change:
twenty rows in which nothing happens demonstrate nothing.

With no price data on disk, a small synthetic series is used and the
sheet says so.
"""
from __future__ import annotations

import argparse
import os
import sys
from collections.abc import Callable
from pathlib import Path
from typing import Any

_HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(_HERE, "roles"))
sys.path.insert(0, os.path.join(_HERE, "..", "..", "..", ".."))

from donchian_signal import (  # noqa: E402
    DONCHIAN_VARIANTS,
    _donchian_compute_variant_signal,
)
from mac_signal import (  # noqa: E402
    MAC_VARIANTS,
    _ewma,
    _mac_compute_variant_signal,
)
from rs_trend import (  # noqa: E402
    RS_TREND_VARIANTS,
    _rs_trend_compute_variant_signal,
)
from turtle_signal import (  # noqa: E402
    TURTLE_VARIANTS,
    _atr_series,
    _turtle_compute_variant_signal,
)

# ── loading bars ───────────────────────────────────────────────────────


#: ``None`` means "wherever market data lives" -- see
#: ``dissyslab.market_data``. This used to be
#: ``_HERE/../../../../sp100_data``, which is the repository root in a
#: clone and nowhere at all after ``dsl init``. The fallback to
#: synthetic prices then hid it: the sheet was produced, it looked
#: right, and only the Read me said the numbers were invented. A
#: graceful fallback is very good at concealing the path everyone
#: actually takes.
DEFAULT_DATA_DIR = None


def _data_dirs(directory):
    try:
        from dissyslab.market_data import search_dirs

        return [str(d) for d in search_dirs(directory, start=Path(_HERE))]
    except ImportError:  # pragma: no cover - standalone use
        return [directory] if directory else []


def load_bars(ticker: str, directory=None, seed: int = 0) -> tuple[list[dict], str]:
    """Real bars if the CSV is there, otherwise a synthetic series.

    Returns ``(bars, provenance)``; the provenance string goes on the
    Read me sheet so nobody mistakes made-up prices for real ones.
    """
    filename = f"{ticker}_10_year.csv"
    searched = _data_dirs(directory)
    for candidate_dir in searched:
        path = os.path.join(candidate_dir, filename)
        if os.path.isfile(path):
            from dissyslab.components.sources.csv_stock_history_source import (
                CSVStockHistorySource,
            )
            src = CSVStockHistorySource(
                tickers=[ticker], directory=candidate_dir,
                filename_pattern="{ticker}_10_year.csv",
            )
            return src._load_ticker(ticker), f"{ticker}, from {path}"

    return _synthetic_bars(seed=seed), (
        f"SYNTHETIC PRICES -- no {filename} in any of: "
        + ", ".join(searched)
        + ". These are made up, so the numbers mean nothing; the "
        "formulas and the rule are still real. Ask your assistant to "
        "download the price history for this ticker, or run "
        "download_stock_history_from_yf.py."
    )


def _synthetic_bars(n: int = 120, seed: int = 0) -> list[dict]:
    """A deterministic sawtooth: rises for 15 bars, falls for 10.

    Chosen so every strategy produces several signal changes in a short
    window, which is what makes a trace worth reading.

    `seed` offsets the phase and tilts the drift, so a synthetic basket
    is a real cross-section rather than five identical stocks. RS ranks
    a stock against its peers, and five identical peers rank by tie-
    break, which would teach a reader something untrue.
    """
    bars, price = [], 100.0 + seed
    up, run = True, 0
    for i in range(n):
        step = (1.4 + 0.15 * seed) if up else (-2.0 + 0.1 * seed)
        price += step
        run += 1
        if (up and run == 15) or (not up and run == 10):
            up, run = not up, 0
        bars.append({
            "date": f"2026-{1 + i // 28:02d}-{1 + i % 28:02d}",
            "open": round(price - step, 2),
            "high": round(price + 0.8, 2),
            "low": round(price - 0.8, 2),
            "close": round(price, 2),
        })
    return bars


# ── traces ─────────────────────────────────────────────────────────────
#
# A trace recomputes the intermediates and then **asserts that they
# imply the signal the real strategy module produced**. If this file
# ever drifts from the strategy it explains, the export fails rather
# than quietly showing the wrong working.
#
# How much of the working fits on the sheet varies, and each tracer
# says which case it is:
#
#   Donchian, MAC  every intermediate is a formula over this ticker's
#                  own rows, so all of it is live and editable.
#   Turtle         the channels and N are formulas; units, the stop and
#                  the signal are a state machine over the whole path,
#                  shown as numbers with the rule that fired.
#   RS             the trend half is a formula; the strength half is
#                  cross-sectional -- computed from the other stocks in
#                  the basket, which are not on this sheet at all.


class Trace:
    def __init__(self, name: str, warmup: int, columns: list[str],
                 rows: list[dict], rule: str, notes: list[str]):
        # Filled in by _trace_sheet once the table length is known.
        self.alpha_cells: dict[str, str] = {}
        self.name = name          # e.g. "donchian_20"
        self.warmup = warmup      # rows of context the formulas need
        self.columns = columns
        self.rows = rows
        self.rule = rule
        self.notes = notes


def donchian_trace(bars: list[dict], variant: str, peers=None) -> Trace:
    n = DONCHIAN_VARIANTS[variant]
    truth = _donchian_compute_variant_signal(bars, n)

    highs = [b["high"] for b in bars]
    lows = [b["low"] for b in bars]
    rows, pos = [], 0.0
    for t, b in enumerate(bars):
        upper = max(highs[t - n:t]) if t >= n else None
        lower = min(lows[t - n:t]) if t >= n else None
        why = "window not full -- flat"
        if t >= n:
            if b["close"] > upper:
                pos, why = 1.0, f"close {b['close']} > upper {upper} -- go long"
            elif b["close"] < lower:
                pos, why = -1.0, f"close {b['close']} < lower {lower} -- go short"
            else:
                why = "no breakout -- hold"
        rows.append({
            "t": t, "date": b["date"], "high": b["high"], "low": b["low"],
            "close": b["close"], "upper": upper, "lower": lower,
            "signal": pos, "why": why,
        })

    _agree(rows, truth, f"donchian_{variant}")
    return Trace(
        name=f"donchian_{variant}", warmup=n,
        columns=["t", "date", "high", "low", "close", "upper", "lower",
                 "signal", "why"],
        rows=rows,
        rule=(f"Donchian({n}). The upper channel is the highest high of "
              f"the {n} bars BEFORE today; the lower channel is the "
              f"lowest low of those same bars. Today's close above the "
              f"upper channel goes long; below the lower channel goes "
              f"short; otherwise hold whatever you held. Flat until the "
              f"first breakout."),
        notes=[f"The first {n} rows are context. The formulas in the "
               f"first visible row read them, so they are shown."],
    )


def mac_trace(bars: list[dict], variant: str, peers=None) -> Trace:
    fast_span, slow_span = MAC_VARIANTS[variant]
    truth = _mac_compute_variant_signal(bars, (fast_span, slow_span))

    closes = [b["close"] for b in bars]
    fast, slow = _ewma(closes, fast_span), _ewma(closes, slow_span)
    rows = []
    for t, b in enumerate(bars):
        sig = 1.0 if fast[t] > slow[t] else -1.0
        if t == 0:
            why = "seeded: both averages equal the first close, so not above -- short"
        elif fast[t] > slow[t]:
            why = f"fast {fast[t]:.4f} > slow {slow[t]:.4f} -- long"
        else:
            why = f"fast {fast[t]:.4f} <= slow {slow[t]:.4f} -- short"
        rows.append({
            "t": t, "date": b["date"], "close": b["close"],
            "fast": fast[t], "slow": slow[t], "signal": sig, "why": why,
        })

    _agree(rows, truth, f"mac_{variant}")
    af, asl = 2.0 / (fast_span + 1), 2.0 / (slow_span + 1)
    return Trace(
        name=f"mac_{variant}", warmup=0,
        columns=["t", "date", "close", "fast", "slow", "signal", "why"],
        rows=rows,
        rule=(f"Exponential moving averages of the close at spans "
              f"{fast_span} and {slow_span} (alpha = 2/(span+1), so "
              f"{af:.4f} and {asl:.4f}). Both are seeded with the first "
              f"close. Long while the faster is above the slower, short "
              f"otherwise -- a tie counts as short."),
        notes=["The averages depend on every earlier bar, so the first "
               "visible row is seeded from the full run rather than "
               "recomputed. Rows below it are the recurrence, each "
               "reading the row above.",
               "A tie goes short. On a flat price series both averages "
               "are equal for ever, so the signal is short for ever. "
               "That follows from the rule as written; it may not be "
               "what was intended."],
    )


def turtle_trace(bars: list[dict], variant: str, peers=None) -> Trace:
    """Turtle is the first strategy here whose position is not a
    formula.

    Donchian and MAC decide today from today's numbers, so every
    intermediate can be shown twice -- once as a number and once as a
    live Excel formula that recomputes it. Turtle carries state: how
    many units are on, what the stop is, what price the last unit was
    added at. Today's position depends on the whole path, not on
    today's row.

    So this sheet shows live formulas for the parts that are formulas
    -- the true range, N, and the entry channels -- and shows units,
    the stop and the signal as numbers with the rule that fired beside
    them. A cell formula for the position would be a lie about how the
    strategy works, and this whole exercise exists so that a person can
    see how it works.
    """
    p = TURTLE_VARIANTS[variant]
    entry_n, exit_n = p["entry"], p["exit"]
    atr_n, max_units, stop_mult = p["atr_period"], p["max_units"], p["stop_atr_mult"]
    unit = 1.0 / max_units
    truth = _turtle_compute_variant_signal(bars, p)

    highs = [b["high"] for b in bars]
    lows = [b["low"] for b in bars]
    closes = [b["close"] for b in bars]
    atr = _atr_series(bars, atr_n)

    rows: list[dict] = []
    units, entry_price, stop_price = 0, None, None

    for t, b in enumerate(bars):
        price = closes[t]
        day_atr = atr[t] if atr[t] else 0.0
        prev_close = closes[t - 1] if t > 0 else closes[t]
        tr = max(b["high"] - b["low"],
                 abs(b["high"] - prev_close),
                 abs(b["low"] - prev_close))

        entry_high = max(highs[t - entry_n:t]) if t >= entry_n else None
        entry_low = min(lows[t - entry_n:t]) if t >= entry_n else None
        exit_low = min(lows[t - exit_n:t]) if t >= exit_n else None
        exit_high = max(highs[t - exit_n:t]) if t >= exit_n else None

        why = "flat -- no breakout yet"
        if units == 0:
            if entry_high is not None and price > entry_high:
                units, entry_price = 1, price
                stop_price = price - stop_mult * day_atr
                why = (f"close {price:.2f} > {entry_n}-day high "
                       f"{entry_high:.2f} -- open 1 unit long, stop "
                       f"{stop_price:.2f}")
            elif entry_low is not None and price < entry_low:
                units, entry_price = -1, price
                stop_price = price + stop_mult * day_atr
                why = (f"close {price:.2f} < {entry_n}-day low "
                       f"{entry_low:.2f} -- open 1 unit short, stop "
                       f"{stop_price:.2f}")
            elif entry_high is not None:
                why = "inside the channel -- stay flat"
        elif units > 0:
            added = False
            if (units < max_units and entry_price is not None and day_atr > 0
                    and price >= entry_price + 0.5 * day_atr):
                units += 1
                entry_price = price
                stop_price = price - stop_mult * day_atr
                added = True
            if (stop_price is not None and price <= stop_price) or (
                    exit_low is not None and price < exit_low):
                reason = ("stopped out" if stop_price is not None
                          and price <= stop_price else
                          f"below the {exit_n}-day low")
                why = f"{reason} at {price:.2f} -- close the position"
                units, entry_price, stop_price = 0, None, None
            elif added:
                why = (f"advanced half an N -- add a unit, now {units} of "
                       f"{max_units}, stop {stop_price:.2f}")
            else:
                why = f"holding {units} unit(s) long, stop {stop_price:.2f}"
        else:
            added = False
            if (units > -max_units and entry_price is not None and day_atr > 0
                    and price <= entry_price - 0.5 * day_atr):
                units -= 1
                entry_price = price
                stop_price = price + stop_mult * day_atr
                added = True
            if (stop_price is not None and price >= stop_price) or (
                    exit_high is not None and price > exit_high):
                reason = ("stopped out" if stop_price is not None
                          and price >= stop_price else
                          f"above the {exit_n}-day high")
                why = f"{reason} at {price:.2f} -- close the position"
                units, entry_price, stop_price = 0, None, None
            elif added:
                why = (f"fell half an N -- add a unit, now {abs(units)} of "
                       f"{max_units} short, stop {stop_price:.2f}")
            else:
                why = f"holding {abs(units)} unit(s) short, stop {stop_price:.2f}"

        rows.append({
            "t": t, "date": b["date"], "high": b["high"], "low": b["low"],
            "close": price, "tr": tr, "N": atr[t],
            "entry_high": entry_high, "entry_low": entry_low,
            "exit_low": exit_low, "exit_high": exit_high,
            "units": units, "stop": stop_price,
            "signal": units * unit, "why": why,
        })

    _agree(rows, truth, f"turtle_{variant}")
    return Trace(
        name=f"turtle_{variant}", warmup=max(entry_n, exit_n, atr_n),
        columns=["t", "date", "high", "low", "close", "tr", "N",
                 "entry_high", "entry_low", "exit_low", "exit_high",
                 "units", "stop", "signal", "why"],
        rows=rows,
        rule=(
            f"Turtle system {variant.upper()}. Enter on a break of the "
            f"{entry_n}-day channel: a close above the highest high of "
            f"the {entry_n} bars BEFORE today goes long, below the "
            f"lowest low goes short. N is the Average True Range over "
            f"{atr_n} days, Wilder-smoothed. Add a unit each time price "
            f"advances half an N in your favour, to {max_units} units, "
            f"resetting the stop to {stop_mult}N from the new unit's "
            f"price. Exit on the stop, or on a close through the "
            f"{exit_n}-day channel the other way. The signal is units / "
            f"{max_units}, so it moves in steps of {unit}."
        ),
        notes=[
            f"The first {max(entry_n, exit_n, atr_n)} rows are context. "
            f"The formulas in the first visible row read them, so they "
            f"are shown.",
            "The channels are built from the rows ABOVE this one, not "
            "including it -- check any shaded cell's formula bar. That "
            "convention is ambiguous in English and decides whether the "
            "backtest was honest.",
            "units, stop and signal are shown as numbers and not as "
            "formulas, on purpose. They depend on the whole path -- what "
            "was open yesterday, where the last unit was added -- so no "
            "cell formula could compute them from this row. The `why` "
            "column is the rule that actually fired.",
            "N uses today's own high, low and yesterday's close, all of "
            "which are known at today's close.",
        ],
    )


#: The basket the shipped office uses. RS is the one strategy whose
#: signal for a ticker depends on the *other* tickers, so tracing it
#: needs peers; without a basket it is flat for ever and the sheet
#: would show nothing.
DEFAULT_PEERS = ["AMD", "NFLX", "NVDA", "PLTR", "TSLA"]


def rs_trace(bars: list[dict], variant: str, peers=None) -> Trace:
    """RS is the first strategy whose working is not all on the sheet.

    Donchian, MAC and Turtle decide a ticker from that ticker. RS asks
    a second question -- is this stock strong *relative to its peers* --
    and the answer comes from the other stocks in the basket, which are
    not on this sheet and cannot be.

    So the percentile is shown as a number, with the basket it came
    from named on the Read me, and only the part computable from this
    ticker's own column gets a live formula: the price `trend_lb` days
    ago, which is what the trend half compares against.

    The percentile is computed by MARKET_CONTEXT's own code, not by a
    second implementation here. Two copies of that arithmetic is how
    the four-dots path bug happened.
    """
    p = RS_TREND_VARIANTS[variant]
    trend_lb, min_pct = p["trend_lb"], p["min_percentile"]

    context = _rs_context(bars, peers)
    truth = _rs_trend_compute_variant_signal(bars, p, context)
    pct = (context or {}).get("rs_percentile") or [None] * len(bars)
    closes = [b["close"] for b in bars]

    rows = []
    for t, b in enumerate(bars):
        ref = closes[t - trend_lb] if t >= trend_lb else None
        trend_up = ref is not None and closes[t] >= ref
        this_pct = pct[t] if t < len(pct) else None
        strong = this_pct is not None and this_pct >= min_pct

        if ref is None:
            why = f"fewer than {trend_lb} bars of history -- flat"
        elif not trend_up:
            why = (f"close {closes[t]:.2f} < the close {trend_lb} days "
                   f"ago ({ref:.2f}) -- not trending, flat")
        elif this_pct is None:
            why = "no peer data for this day -- flat"
        elif not strong:
            why = (f"trending, but only {this_pct:.2f} of peers are "
                   f"weaker (needs {min_pct:.2f}) -- flat")
        else:
            why = (f"trending and stronger than {this_pct:.2f} of peers "
                   f"-- hold a full position")

        rows.append({
            "t": t, "date": b["date"], "close": closes[t],
            "trend_ref": ref, "trending": trend_up,
            "rs_percentile": this_pct, "strong": strong,
            "signal": 1.0 if (trend_up and strong) else 0.0, "why": why,
        })

    _agree(rows, truth, f"rs_{variant}")
    return Trace(
        name=f"rs_{variant}", warmup=trend_lb,
        columns=["t", "date", "close", "trend_ref", "trending",
                 "rs_percentile", "strong", "signal", "why"],
        rows=rows,
        rule=(
            f"Long only, and only when both halves agree. The trend "
            f"half: today's close at or above the close {trend_lb} "
            f"trading days ago. The strength half: this stock's "
            f"{trend_lb}-day momentum ranks in the top "
            f"{100 * (1 - min_pct):.0f}% of the basket that day -- "
            f"rs_percentile is the fraction of peers that were weaker, "
            f"so it must be at least {min_pct}. Either half failing "
            f"means flat, holding nothing."
        ),
        notes=[
            f"The first {trend_lb} rows are context -- the trend "
            f"comparison reads them.",
            "rs_percentile is cross-sectional: it comes from the other "
            "stocks in the basket on that date, which are not on this "
            "sheet. It has no Excel formula beside it for that reason, "
            "and the Read me names the basket it was computed from. "
            "Change the basket and this column changes.",
            "This is the only shipped strategy that spends whole "
            "stretches out of the market, which is why the report has "
            "a days-in-market column.",
        ],
    )


def _rs_context(bars: list[dict], peers) -> dict:
    """This ticker's causal relative-strength series, from
    MARKET_CONTEXT's own implementation.

    `peers` is ``{ticker: bars}`` for the whole basket, this ticker
    included. With no peers there is no cross-section, and the
    strategy is flat for ever -- which is what the role does too.
    """
    if not peers:
        return {}
    from market_context import make_market_context

    out = make_market_context()({"history": peers})[0][0]
    context = out.get("context") or {}
    this = next((t for t, b in peers.items() if b is bars), None)
    if this is None:
        return {}
    per = (context.get("per_ticker") or {}).get(this) or {}
    by_date = per.get("rs_percentile_by_date", {})
    return {"rs_percentile": [by_date.get(b["date"]) for b in bars]}


def _agree(rows: list[dict], truth: list[float], label: str) -> None:
    mine = [r["signal"] for r in rows]
    if mine != list(truth):
        bad = next(i for i, (a, b) in enumerate(zip(mine, truth)) if a != b)
        raise SystemExit(
            f"{label}: this explainer disagrees with the strategy at bar "
            f"{bad} (explainer {mine[bad]}, strategy {truth[bad]}). The "
            f"working shown here would be wrong, so nothing was written. "
            f"Fix explain_strategy.py to match the role."
        )


TRACERS: dict[str, tuple[Callable, dict[str, Any]]] = {
    "donchian": (donchian_trace, DONCHIAN_VARIANTS),
    "mac": (mac_trace, MAC_VARIANTS),
    "turtle": (turtle_trace, TURTLE_VARIANTS),
    "rs": (rs_trace, RS_TREND_VARIANTS),
}


# ── choosing the window ────────────────────────────────────────────────


def pick_window(trace: Trace, rows: int) -> tuple[int, int]:
    """A window containing a signal change, or the last `rows` bars.

    Twenty bars in which nothing happens prove nothing, so the default
    centres on the first change after warmup.
    """
    sig = [r["signal"] for r in trace.rows]
    change = next(
        (i for i in range(trace.warmup + 1, len(sig)) if sig[i] != sig[i - 1]),
        None,
    )
    if change is None:
        return max(0, len(sig) - rows), len(sig)
    start = max(0, change - rows // 3)
    return start, min(len(sig), start + rows)


# ── writing the workbook ───────────────────────────────────────────────


HEADER_FILL = "DDEBF7"
FORMULA_FILL = "FFF2CC"
CONTEXT_FILL = "F2F2F2"


def write_workbook(traces: list[Trace], windows: list[tuple[int, int]],
                   provenance: str, out_path: str) -> None:
    try:
        import openpyxl
    except ImportError:
        raise SystemExit(
            "This needs openpyxl to write the spreadsheet, and it is not "
            "installed.\n\n"
            '    pip install "dissyslab[market]"\n\n'
            "That extra carries yfinance for the price history and openpyxl "
            "for this. Use the same Python that runs `dsl` -- if dsl lives "
            "in a virtualenv, activate it first."
        ) from None

    wb = openpyxl.Workbook()
    # Nothing here computes a formula, so every formula cell ships with
    # no stored answer. Excel is entitled to trust a stored answer if
    # one is present; asking for a full recalculation on load means the
    # numbers a reader sees were worked out by their spreadsheet from
    # the prices in front of them, and not by us. It also fills the
    # columns in immediately, which is what stops the sheet looking
    # broken on first open.
    wb.calculation.fullCalcOnLoad = True
    _read_me(wb.active, traces, provenance)
    for tr, (lo, hi) in zip(traces, windows):
        _trace_sheet(wb.create_sheet(tr.name[:31]), tr, lo, hi)
    wb.save(out_path)


def _read_me(ws, traces: list[Trace], provenance: str) -> None:
    from openpyxl.styles import Alignment, Font

    ws.title = "Read me"
    lines: list[tuple[str, bool]] = [
        ("What this is", True),
        ("One row per trading day, showing every quantity the strategy "
         "computed and the rule that turned them into a position.", False),
        ("", False),
        ("Shaded columns are live Excel formulas, not numbers. Click one "
         "and read the formula bar: it is the rule, in a form you can "
         "check and change. Change a price and the shaded columns "
         "recompute.", False),
        ("", False),
        ("Open this in a spreadsheet application -- Excel, Numbers, "
         "Google Sheets. A formula has no stored answer until something "
         "works it out, so in a preview that does not calculate (the "
         "Finder's Quick Look, a file listing on the web) the shaded "
         "columns and the match columns look empty. They are not empty; "
         "nothing has evaluated them yet.", False),
        ("", False),
        ("The unshaded value columns are what the Python produced. They "
         "do not verify the formulas -- both say the same author's "
         "understanding of the rule. If they disagree, one of the two "
         "has a slip and the match column says so.", False),
        ("", False),
        ("Prices", True),
        (provenance, False),
        ("", False),
    ]
    for tr in traces:
        lines.append((f"Sheet: {tr.name}", True))
        lines.append((tr.rule, False))
        for note in tr.notes:
            lines.append((f"  - {note}", False))
        lines.append(("", False))

    for i, (text, bold) in enumerate(lines, start=1):
        c = ws.cell(row=i, column=1, value=text)
        c.font = Font(name="Arial", bold=bold, size=12 if bold else 11)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 100


def _trace_sheet(ws, tr: Trace, lo: int, hi: int) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    first = max(0, lo - tr.warmup)          # context the formulas need
    shown = tr.rows[first:hi]

    headers = _headers(tr)
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(name="Arial", bold=True)
        c.fill = PatternFill("solid", fgColor=HEADER_FILL)
        c.alignment = Alignment(wrap_text=True, vertical="top")

    col = {name: chr(ord("A") + i) for i, name in enumerate(headers)}
    anchor = len(shown) + 3
    tr.alpha_cells = {"fast": f"$B${anchor}", "slow": f"$B${anchor + 1}"}
    for i, r in enumerate(shown):
        excel_row = i + 2
        ws.append(_row_values(tr, r, shown, i, excel_row, col))

    _style(ws, tr, shown, lo, first, col, headers)
    ws.freeze_panes = "A2"


def _headers(tr: Trace) -> list[str]:
    out: list[str] = []
    for name in tr.columns:
        out.append(name)
        if name in _DERIVED[tr.name.split("_")[0]]:
            out.append(f"{name} (excel)")
            out.append("match")
    return out


_DERIVED = {
    "donchian": ("upper", "lower"),
    "mac": ("fast", "slow"),
    # Turtle's position is a state machine, not a formula -- only the
    # quantities that *are* formulas get a live Excel column beside
    # them. See turtle_trace's docstring.
    "turtle": ("tr", "N", "entry_high", "entry_low"),
    # RS's percentile is cross-sectional -- it is computed from the
    # other stocks in the basket, which are not on this sheet. Only
    # the part computable from this ticker's own column gets a live
    # formula. See rs_trace's docstring.
    "rs": ("trend_ref",),
}


def _row_values(tr: Trace, r: dict, shown: list[dict], i: int,
                excel_row: int, col: dict[str, str]) -> list[Any]:
    kind = tr.name.split("_")[0]
    out: list[Any] = []
    for name in tr.columns:
        out.append(r.get(name))
        if name not in _DERIVED[kind]:
            continue
        formula = _formula(tr, kind, name, r, i, excel_row, col)
        out.append(formula)
        vcol, fcol = col[name], col[f"{name} (excel)"]
        # Compare only where there are two things to compare. A context
        # row can carry a value and no formula -- the formula would
        # need rows above the top of the sheet -- and comparing a
        # number against an empty cell reported DIFFERS for every one
        # of them. Found by recalculating an rs sheet in LibreOffice
        # and reading this column, which is what it is for.
        out.append(
            f'=IF(OR({vcol}{excel_row}="",{fcol}{excel_row}=""),"",'
            f'IF(ROUND({vcol}{excel_row},6)=ROUND({fcol}{excel_row},6),'
            f'"ok","DIFFERS"))'
        )
    return out


def _peers_for(strategy: str, args, bars: list[dict]) -> dict | None:
    """The basket, loaded only for the strategy that needs it.

    RS compares a stock against its peers, so tracing it means loading
    them. Everything else is a per-ticker rule and pays nothing for
    this.
    """
    if strategy != "rs":
        return None
    names = [t.strip() for t in (args.peers or ",".join(DEFAULT_PEERS)).split(",")]
    peers = {args.ticker: bars}
    for name in names:
        if name == args.ticker:
            continue
        other, _prov = load_bars(name, args.data_dir, seed=abs(hash(name)) % 7)
        peers[name] = other
    return peers


def _formula(tr: Trace, kind: str, name: str, r: dict, i: int,
             excel_row: int, col: dict[str, str]) -> str | None:
    if kind == "rs":
        # Only the trend half is on this sheet. The percentile comes
        # from the other stocks in the basket and has no formula here.
        if r[name] is None or i < tr.warmup:
            return None
        return f"={col['close']}{excel_row - tr.warmup}"

    if kind == "donchian":
        if r[name] is None or i < tr.warmup:
            return None
        src = col["high"] if name == "upper" else col["low"]
        fn = "MAX" if name == "upper" else "MIN"
        return (f"={fn}({src}{excel_row - tr.warmup}:{src}{excel_row - 1})")

    if kind == "turtle":
        return _turtle_formula(tr, name, r, i, excel_row, col)

    # mac: the recurrence, seeded on the first visible row because the
    # average depends on every earlier bar.
    if i == 0:
        return r[name]
    alpha_cell = tr.alpha_cells[name]
    return (f"={alpha_cell}*{col['close']}{excel_row}"
            f"+(1-{alpha_cell})*{col[name + ' (excel)']}{excel_row - 1}")


def _turtle_formula(tr: Trace, name: str, r: dict, i: int,
                    excel_row: int, col: dict[str, str]) -> str | None:
    """Only the quantities that really are formulas get one.

    `units`, `stop` and `signal` are not here on purpose: they depend
    on the whole path, and a cell formula pretending otherwise would
    misrepresent the strategy to the one person reading the sheet to
    find out how it works.
    """
    p = TURTLE_VARIANTS[tr.name.split("_", 1)[1]]
    high, low, close = col["high"], col["low"], col["close"]

    if name == "tr":
        if i == 0:
            return r[name]        # no row above to read yesterday's close
        prev = excel_row - 1
        return (f"=MAX({high}{excel_row}-{low}{excel_row},"
                f"ABS({high}{excel_row}-{close}{prev}),"
                f"ABS({low}{excel_row}-{close}{prev}))")

    if name == "N":
        n = p["atr_period"]
        # Two regimes, and getting this wrong is what the match column
        # is for -- it caught exactly this. Before the window has
        # filled, `_atr_series` uses a plain mean of the true ranges so
        # far, not the recurrence; writing the recurrence for those
        # rows made every later row disagree, because a recurrence
        # carries its seed forward for ever. Those rows also read bars
        # that may be off the top of the sheet, so they are shown as
        # values.
        if i == 0 or r["t"] < n:
            return r[name]
        # Wilder smoothing proper, each row reading the row above.
        return (f"=({col['N (excel)']}{excel_row - 1}*{n - 1}"
                f"+{col['tr (excel)']}{excel_row})/{n}")

    if r[name] is None or i < tr.warmup:
        return None
    entry = p["entry"]
    src = high if name == "entry_high" else low
    fn = "MAX" if name == "entry_high" else "MIN"
    return f"={fn}({src}{excel_row - entry}:{src}{excel_row - 1})"


def _style(ws, tr: Trace, shown: list[dict], lo: int, first: int,
           col: dict[str, str], headers: list[str]) -> None:
    from openpyxl.styles import Font, PatternFill

    formula_cols = [c for h, c in col.items() if h.endswith("(excel)")]
    for i in range(len(shown)):
        row = i + 2
        for c in ws[row]:
            c.font = Font(name="Arial")
        for letter in formula_cols:
            ws[f"{letter}{row}"].fill = PatternFill("solid", fgColor=FORMULA_FILL)
        if first + i < lo:                      # context rows
            for c in ws[row]:
                if c.column_letter not in formula_cols:
                    c.fill = PatternFill("solid", fgColor=CONTEXT_FILL)

    if tr.name.startswith("mac"):
        # Below the table, not beside it -- the header row is occupied,
        # and writing into it silently replaced the 'why' heading.
        fast_span, slow_span = MAC_VARIANTS[tr.name.split("_", 1)[1]]
        anchor = len(shown) + 3
        ws[f"A{anchor}"] = "alpha fast = 2/(span+1)"
        ws[f"A{anchor + 1}"] = "alpha slow = 2/(span+1)"
        ws[f"B{anchor}"] = 2.0 / (fast_span + 1)
        ws[f"B{anchor + 1}"] = 2.0 / (slow_span + 1)
        for r in (anchor, anchor + 1):
            ws[f"A{r}"].font = Font(name="Arial", bold=True)
            ws[f"B{r}"].font = Font(name="Arial")
            ws[f"B{r}"].number_format = "0.000000"

    widths = {"t": 5, "date": 11, "why": 46}
    for h, letter in col.items():
        ws.column_dimensions[letter].width = widths.get(h, 13)
    for row in ws.iter_rows(min_row=2):
        for c in row:
            if isinstance(c.value, float):
                c.number_format = "0.0000"


# ── entry point ────────────────────────────────────────────────────────


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(description=__doc__.split("\n")[0])
    p.add_argument("--strategy", action="append", default=None,
                   choices=sorted(TRACERS), help="repeatable")
    p.add_argument("--variant", action="append", default=None,
                   help="repeatable; defaults to the first of each strategy")
    p.add_argument("--ticker", default="AMD")
    p.add_argument(
        "--peers",
        help="comma-separated basket for rs, which ranks a stock "
             "against its peers. Defaults to the shipped office's "
             "basket. Ignored by every other strategy.",
    )
    p.add_argument("--data-dir", default=DEFAULT_DATA_DIR)
    p.add_argument("--bars", default=None, metavar="LO:HI",
                   help="bar range; default is a window containing a change")
    p.add_argument("--rows", type=int, default=30)
    p.add_argument("--out", default="strategy_working.xlsx")
    a = p.parse_args(argv)

    strategies = a.strategy or ["donchian"]
    # The seed only matters when there is no real data. It is
    # derived from the name so that a synthetic basket has this
    # ticker somewhere inside it rather than always at the bottom:
    # rs ranks against peers, and a stock that is weakest by
    # construction is flat on every row.
    bars, provenance = load_bars(
        a.ticker, a.data_dir, seed=abs(hash(a.ticker)) % 7)

    traces, windows = [], []
    for i, s in enumerate(strategies):
        tracer, variants = TRACERS[s]
        variant = (a.variant[i] if a.variant and i < len(a.variant)
                   else sorted(variants)[0])
        if variant not in variants:
            p.error(f"{s}: unknown variant {variant!r}; "
                    f"choose from {sorted(variants)}")
        tr = tracer(bars, variant, peers=_peers_for(s, a, bars))
        traces.append(tr)
        if a.bars:
            lo, hi = (int(x) for x in a.bars.split(":"))
        else:
            lo, hi = pick_window(tr, a.rows)
        windows.append((lo, hi))

    write_workbook(traces, windows, provenance, a.out)
    print(f"wrote {a.out}")
    for tr, (lo, hi) in zip(traces, windows):
        print(f"  {tr.name}: bars {lo}-{hi - 1}"
              f"{f' (+{tr.warmup} rows of context)' if tr.warmup else ''}")
    print("Formula cells have no cached value until a spreadsheet "
          "application opens the file.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
